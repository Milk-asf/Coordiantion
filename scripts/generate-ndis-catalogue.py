#!/usr/bin/env python3
"""Generate src/lib/data/ndis-pricing-catalogue.json from the NDIS Support Catalogue XLSX."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    raise

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "src/lib/data/ndis-pricing-catalogue.json"

UNIT_MAP = {
    "H": "hour",
    "E": "each",
    "D": "day",
    "WK": "week",
    "MON": "month",
    "YR": "year",
}

STATE_PRICE_COLUMNS = ["NSW", "VIC", "QLD", "SA", "WA", "TAS", "ACT", "NT", "Remote", "Very Remote"]


def slugify(value: str) -> str:
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", value.lower())).strip("-")


def make_short_name(name: str) -> str:
    if not name:
        return ""
    if ":" in name:
        left, right = name.split(":", 1)
        candidate = right.strip() if len(right.strip()) <= 40 else left.strip()
        return candidate[:40]
    return name[:40]


def pick_price(row: dict) -> float:
    for col in STATE_PRICE_COLUMNS:
        val = row.get(col)
        if isinstance(val, (int, float)) and val > 0:
            return float(val)
    return 0.0


def main() -> None:
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/tmp/ndis-catalogue.xlsx")
    if not source.exists():
        print(f"Missing source file: {source}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    ws = wb["Current Support Items"]
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

    items: list[dict] = []
    categories: dict[str, str] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        item_number = str(data.get("Support Item Number") or "").strip()
        name = str(data.get("Support Item Name") or "").strip()
        if not item_number or not name:
            continue

        support_category = str(
            data.get("Support Category Name (PACE)") or data.get("Support Category Name") or ""
        ).strip()
        registration_group = str(data.get("Registration Group Name") or "").strip()
        registration_group_number = str(data.get("Registration Group Number") or "").strip()
        support_category_number = int(
            data.get("Support Category Number (PACE)") or data.get("Support Category Number") or 0
        )
        unit_code = str(data.get("Unit") or "E").strip().upper()
        unit = UNIT_MAP.get(unit_code, "each")
        quote_required = str(data.get("Quote") or "").strip().lower() == "yes"
        price = pick_price(data)
        category_id = slugify(support_category) or "other"

        categories[category_id] = support_category or "Other"

        items.append(
            {
                "itemNumber": item_number,
                "name": name,
                "shortName": make_short_name(name),
                "registrationGroup": registration_group,
                "registrationGroupNumber": registration_group_number,
                "supportCategory": support_category,
                "supportCategoryNumber": support_category_number,
                "category": category_id,
                "unit": unit,
                "price": price,
                "quoteRequired": quote_required,
            }
        )

    supplemental = {
        "itemNumber": "07_799_0106_6_3_KM",
        "name": "Provider Travel — Kilometres",
        "shortName": "Travel (km)",
        "registrationGroup": "Support Coordination",
        "registrationGroupNumber": "0106",
        "supportCategory": "Support Coordination and Psychosocial Recovery Coaches",
        "supportCategoryNumber": 7,
        "category": "support-coordination-and-psychosocial-recovery-coaches",
        "unit": "km",
        "price": 0.99,
        "quoteRequired": False,
    }
    if not any(item["itemNumber"] == supplemental["itemNumber"] for item in items):
        items.append(supplemental)

    payload = {
        "version": "2025-26",
        "effectiveFrom": "2024-07-01",
        "source": "NDIS Support Catalogue (Current Support Items)",
        "sourceNote": "Generated from the official NDIS Support Catalogue spreadsheet. NSW price used when available; quote-required items have no set price.",
        "itemCount": len(items),
        "categories": dict(sorted(categories.items(), key=lambda kv: kv[1])),
        "items": sorted(items, key=lambda item: (item["supportCategoryNumber"], item["name"])),
    }

    OUT.write_text(json.dumps(payload, indent=2))
    print(f"Wrote {len(items)} items to {OUT}")


if __name__ == "__main__":
    main()
